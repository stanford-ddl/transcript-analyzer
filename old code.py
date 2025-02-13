import os
import json
import asyncio
import aiohttp
from datetime import datetime, timezone
import keyring
from openpyxl import load_workbook
from tkinter import Tk, filedialog
from typing import Dict
import logging

# Constants for Anthropic API
API_URL = "https://api.anthropic.com/v1/messages"
API_KEY = keyring.get_password("Anthropic_personal", "Metaverse transcripts")  # Replace with your actual API key

# Debug logging
DEBUG = True
LOG_FILE = f"/Users/rickyhm/Onboard/DDL-Transcript-Analyzer-1.1.2/rickys_version/debug logs/debug_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

# Configure logging
logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG if DEBUG else logging.INFO)

# Global token counters
total_input_tokens = 0
total_output_tokens = 0

# Rate limit state
rate_limit_state = {
    "requests_remaining": float("inf"),
    "tokens_remaining": float("inf"),
    "reset_time": None,
}

def log_debug_message(message: str):
    """
    Log a debug message to the log file and optionally print it.
    """
    logging.debug(message)
    if DEBUG:
        print(message)

def parse_reset_time(reset_time_str):
    """
    Parse RFC 3339 reset time string into a datetime object.
    """
    return datetime.fromisoformat(reset_time_str.replace("Z", "+00:00")).replace(tzinfo=timezone.utc)

def handle_rate_limiting(headers):
    """
    Update rate limit state from API response headers.
    """
    global rate_limit_state

    rate_limit_state["requests_remaining"] = int(headers.get("anthropic-ratelimit-requests-remaining", float("inf")))
    rate_limit_state["tokens_remaining"] = int(headers.get("anthropic-ratelimit-tokens-remaining", float("inf")))
    reset_time_str = headers.get("anthropic-ratelimit-requests-reset")

    if reset_time_str:
        rate_limit_state["reset_time"] = parse_reset_time(reset_time_str)

    # Log rate limit details
    log_debug_message(f"[DEBUG] Updated rate limit state: {rate_limit_state}")

async def wait_for_reset():
    """
    Wait until the rate limit resets if necessary.
    """
    if rate_limit_state["reset_time"]:
        now = datetime.now(timezone.utc)
        sleep_time = (rate_limit_state["reset_time"] - now).total_seconds()
        if sleep_time > 0:
            log_debug_message(f"[INFO] Rate limit hit. Sleeping for {sleep_time:.2f} seconds.")
            await asyncio.sleep(sleep_time)

async def make_request(payload):
    """
    Make a request to the Anthropic API and handle rate limits.
    """
    while True:
        if rate_limit_state["requests_remaining"] <= 0 or rate_limit_state["tokens_remaining"] <= 0:
            await wait_for_reset()

        headers = {
            "x-api-key": API_KEY,
            "content-type": "application/json",
            "anthropic-version": "2023-06-01",
        }

        async with aiohttp.ClientSession() as session:
            async with session.post(API_URL, headers=headers, json=payload) as response:
                if response.status == 429:  # Rate limit exceeded
                    log_debug_message("[WARN] Rate limit exceeded. Waiting for reset.")
                    handle_rate_limiting(response.headers)
                    await wait_for_reset()
                    continue

                if response.status != 200:
                    text = await response.text()
                    log_debug_message(f"[ERROR] API request failed: {response.status} - {text}")
                    raise Exception(f"API request failed with status {response.status}")

                handle_rate_limiting(response.headers)
                response_json = await response.json()
                return response_json

async def analyze_utterance(text: str, proposals: dict) -> dict:
    """
    Analyze a single utterance using Anthropic's API via HTTP.
    """
    global total_input_tokens, total_output_tokens

    proposals_text = "\n".join([f"{proposal}" for proposal in proposals.items()])

    # Construct the prompt
    prompt = f"""
You are analyzing transcript excerpts from a discussion about the Metaverse (online virtual reality spaces).

Your task is to determine if the provided text contains any arguments directly stated by the speaker related to the following proposals. Rewrite any argument you find into a single coherent sentence using this structure:

"[Claim] because [reasons]."

### Proposals

Video capture should be used in members-only spaces and/or public spaces.
Automatic speech detection should be used in members-only spaces and/or public spaces.
Creators should be responsible for managing bad behavior in members-only spaces and/or public spaces.
Platform owners should be responsible for managing bad behavior in members-only spaces and/or public spaces.
Punishments should be administered for bad behavior in the Metaverse/online virtual reality spaces.

### Definitions and Criteria

#### What is an Argument?

An argument is a statement explicitly made by the speaker that:
1. Relates directly to one of the proposals.
2. Expresses a clear position (for or against) about the proposal.
3. Provides a justification or reason supporting this position.

#### Important Notes:
- Only consider arguments that the speaker **directly states or clearly implies**. Do not infer beliefs, intentions, or unstated positions.
- If the speaker does not explicitly state a position or reason, do not rewrite the text as an argument, even if the position could be inferred.
- The rewritten argument must be tied to the exact words and intent of the speaker as presented in the text.

#### Sentence Structure for Rewritten Arguments:
- Always rewrite the argument as a single sentence following this format:
  - **"[Claim] because [reason]."**

If the text does not provide a reason (explicitly or implicitly), do not rewrite it as an argument.

### Output Format

If the text contains an argument with one or multiple reasons, rewrite it as a single sentence in the following format:
"Video capture in public spaces should be used because it helps deter crime."

If the text contains multiple arguments for different proposals, rewrite each argument as it's own argument sentence and claim pair, grouping them using xml tags.

<argument_sentence>

If no arguments are present, return "None". Do not include any other text or explanation.

Only return the json (with regular brackets) in your response. No other explanation or text.

Key Points:
	1.	The rewritten argument must include both a claim and a reason, explicitly connected with “because.”
	2.	Only include arguments that are explicitly stated or clearly implied in the speaker’s words.
	3.	Do not infer unstated beliefs or positions from the text.
	4.	If the text contains multiple arguments for different claims, include each as a separate rewritten sentence.

Text:
{text}
"""

    payload = {
        "model": "claude-3-5-sonnet-20241022",
        "max_tokens": 2500,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "<examples>\n<example>\n<text>\nVideo capture should be used in public spaces because it helps deter crime\n</text>\n<ideal_output>\nVideo capture in public spaces should be used because it helps deter crime.\n</ideal_output>\n</example>\n<example>\n<text>\nWe need to protect people in public spaces, so I think video capture should be used.\n</text>\n<ideal_output>\nVideo capture in public spaces should be used because it helps protect people.\n</ideal_output>\n</example>\n<example>\n<example_description>\nNo argument provided because there is no justification\n</example_description>\n<text>\nI think it should be enabled.\n</text>\n<ideal_output>\nNone\n</ideal_output>\n</example>\n<example>\n<example_description>\nThe speaker does not explicitly take a position for or against the proposal and does not provide a clear justification\n</example_description>\n<text>\nVideo capture can be useful, but it depends on how it is used.\n</text>\n<ideal_output>\nNone\n</ideal_output>\n</example>\n<example>\n<text>\nI don’t think we need video capture in public spaces since it’s an invasion of privacy.\n</text>\n<ideal_output>\nVideo capture in public spaces should not be used because it invades privacy.\n</ideal_output>\n</example>\n</examples>\n\n"
                    },
                    {
                        "type": "text",
                        "text": prompt
                    }
                ]
            }
        ]
    }

    # Make the HTTP request
    response = await make_request(payload)

    # Extract token usage
    usage = response.get("usage", {})
    input_tokens = usage.get("input_tokens", 0)
    output_tokens = usage.get("output_tokens", 0)
    total_input_tokens += input_tokens
    total_output_tokens += output_tokens

    log_debug_message(f"[DEBUG] Input tokens: {input_tokens}, Output tokens: {output_tokens}")
    log_debug_message(f"[DEBUG] LLM input:\n{prompt}")

    # Extract the single text field from content
    content = response.get("content", [])
    if not content:
        return {"response_text": "No content in response."}

    # Directly fetch the text field of the first content object
    response_text = content[0].get("text", "No text in content.")
    log_debug_message(f"[DEBUG] LLM output (raw):\n{response_text}")

    return format_llm_response(response_text)

def load_proposals(proposal_file: str) -> Dict[str, str]:
    """
    Load proposals from an Excel file.
    """
    wb = load_workbook(proposal_file)
    proposal_ws = wb["proposals"]
    proposals = {}
    for row in range(2, proposal_ws.max_row + 1):
        key = proposal_ws.cell(row=row, column=1).value
        proposal = proposal_ws.cell(row=row, column=2).value
        if key and proposal:
            proposals[key] = proposal

    if DEBUG:
        print(f"[DEBUG] Loaded proposals: {proposals}")
    return proposals

def format_llm_response(response_data: dict) -> dict:
    """Return the raw LLM response."""
    return response_data

async def process_utterance(utterance: dict, proposals: dict, semaphore: asyncio.Semaphore, idx: int) -> dict:
    """
    Process a single utterance.
    """
    async with semaphore:
        text = utterance.get("text", "")
        result = await analyze_utterance(text, proposals)

    # Parse result if it's a string
    if isinstance(result, str):
        try:
            result = json.loads(result)
        except json.JSONDecodeError as e:
            print(f"[ERROR] Failed to parse result as JSON: {e}")
            result = {}

    # Validate result type
    if isinstance(result, dict):
        if result.get("arguments") is not None:
            utterance["arguments"] = result["arguments"]
        else:
            utterance["arguments"] = []
    else:
        print(f"[ERROR] Unexpected result type: {type(result)}")
        utterance["arguments"] = []
        print(f"Processed utterance {idx}")
        return utterance

async def process_transcript(json_file_path: str, processed_folder: str, proposals: dict, max_concurrent_utterances: int):
    """
    Process a single transcript JSON file.
    """
    with open(json_file_path, 'r') as f:
        data = json.load(f)

    file_name = data.get("filename", os.path.basename(json_file_path))
    utterances = data.get("utterances", [])

    semaphore = asyncio.Semaphore(max_concurrent_utterances)

    tasks = [
        process_utterance(utterance, proposals, semaphore, idx)
        for idx, utterance in enumerate(utterances, start=1)
    ]

    processed_utterances = await asyncio.gather(*tasks)

    # Save processed transcript
    processed_data = {
        "file_name": file_name,
        "utterances": processed_utterances
    }

    output_path = os.path.join(processed_folder, f"{os.path.splitext(file_name)[0]}_processed.json")
    with open(output_path, 'w') as f:
        json.dump(processed_data, f, indent=2)

    if DEBUG:
        log_debug_message(f"[DEBUG] Processed transcript saved to {output_path}")

async def process_all_transcripts(json_folder: str, processed_folder: str, proposals: Dict[str, str], max_concurrent_utterances: int):
    """
    Process all transcript JSON files in the folder (including subfolders).
    """
    for root, _, files in os.walk(json_folder):
        for file_name in files:
            if file_name.endswith(".json"):
                json_file_path = os.path.join(root, file_name)

                # Check if the file already exists in the output folder
                processed_file_path = os.path.join(processed_folder, f"{os.path.splitext(file_name)[0]}_processed.json")
                if os.path.exists(processed_file_path):
                    if DEBUG:
                        print(f"[DEBUG] Skipping {file_name}, already processed.")
                    continue  # Skip processing this file

                if DEBUG:
                    print(f"[DEBUG] Processing transcript: {json_file_path}")

                await process_transcript(json_file_path, processed_folder, proposals, max_concurrent_utterances)

# File and folder selection functions
def select_input():
    """
    Brings up a file dialog for the user to select a folder or file for input.
    """
    root = Tk()
    root.withdraw()  # Hide the main tkinter window
    root.update()
    input_path = filedialog.askopenfilename(title="Select a file or folder for input")
    if not input_path:  # If no file is selected, allow folder selection
        input_path = filedialog.askdirectory(title="Select a folder for input")
    root.destroy()
    return input_path

def select_output():
    """
    Brings up a file dialog for the user to select a folder for output.
    """
    root = Tk()
    root.withdraw()  # Hide the main tkinter window
    root.update()
    output_path = filedialog.askdirectory(title="Select a folder for output")
    root.destroy()
    return output_path

async def main():
    """
    Main async function to handle the script execution.
    """
    global total_input_tokens, total_output_tokens

    # Input and output folders
    print("Please select the input folder or file.")
    json_folder = select_input()
    print(f"Input selected: {json_folder}")

    print("Please select the output folder.")
    processed_folder = select_output()
    print(f"Output selected: {processed_folder}")

    proposal_file = '/Users/rickyhm/Onboard/DDL-Transcript-Analyzer-1.1.2/rickys_version/proposals copy.xlsx'  # Update this path accordingly

    # Load proposals
    proposals = load_proposals(proposal_file)

    # Set the maximum number of concurrent utterances
    max_concurrent_utterances = 50

    if os.path.isfile(json_folder):
        print(f"Processing single file: {json_folder}")
        await process_transcript(json_folder, processed_folder, proposals, max_concurrent_utterances)
    elif os.path.isdir(json_folder):
        print(f"Processing all transcripts in directory: {json_folder}")
        await process_all_transcripts(json_folder, processed_folder, proposals, max_concurrent_utterances)
    else:
        raise ValueError(f"Invalid input path: {json_folder}")

    # Log final token usage
    if DEBUG:
        log_debug_message(f"[DEBUG] Total input tokens used in session: {total_input_tokens}")
        log_debug_message(f"[DEBUG] Total output tokens used in session: {total_output_tokens}")

if __name__ == "__main__":
    # Run the async main function
    asyncio.run(main())
